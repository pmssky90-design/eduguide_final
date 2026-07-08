from __future__ import annotations

import hashlib
import json
import re
import shutil
from collections import Counter, defaultdict
from dataclasses import dataclass
from html import escape
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
OUTPUT_DIR = ROOT / "output"
STATIC_DIR = ROOT / "static"

REGION_XLSX = Path("C:/gptwp") / "자료" / "울산 전주 광주 주요키워드.xlsx"
CONTENT_XLSX = Path("C:/gptwp") / "자료" / "울산_전주_광주_12개시트.xlsx"
EXPECTED_CONTENT_ROWS = 1620
SITE_NAME = "에듀가이드"
BASE_URL = "https://eduguide.kr"
SITE_URL = BASE_URL
FIXED_IMAGE_DIR = ROOT / "images" / "thumbs" / "fixed"
FIXED_IMAGE_URL_BASE = f"{BASE_URL}/images/thumbs/fixed"
FIXED_IMAGE_SRC_BASE = "/images/thumbs/fixed"
FIXED_IMAGE_NAMES = tuple(f"{index:03}.webp" for index in range(1, 7))
FIXED_IMAGE_WIDTH = 1024
FIXED_IMAGE_HEIGHT = 1536
HERO_IMAGE_DIR = ROOT / "images" / "hero"
HERO_IMAGE_NAME = "main-hero.webp"
HERO_IMAGE_WIDTH = 1717
HERO_IMAGE_HEIGHT = 916
MIN_TEXT_LENGTH = 1200
RELATED_LIMIT = 12
RELATED_SECTION_LIMIT = 8
RELATED_TOTAL_LIMIT = 24
CONTENT_SUFFIXES = [
    "과외",
    "수학과외",
    "영어과외",
    "초등과외",
    "중등과외",
    "고등과외",
    "초등수학과외",
    "중등수학과외",
    "고등수학과외",
    "초등영어과외",
    "중등영어과외",
    "고등영어과외",
]
CONTENT_SUFFIX_PRIORITY = [
    "과외",
    "영어과외",
    "수학과외",
    "초등과외",
    "중등과외",
    "고등과외",
    "초등영어과외",
    "중등영어과외",
    "고등영어과외",
    "초등수학과외",
    "중등수학과외",
    "고등수학과외",
]
BASE_SUFFIX = CONTENT_SUFFIXES[0]
MATH_SUFFIX = CONTENT_SUFFIXES[1]
ENGLISH_SUFFIX = CONTENT_SUFFIXES[2]
SUBJECT_SUFFIXES = [MATH_SUFFIX, ENGLISH_SUFFIX]
GRADE_SUFFIXES = CONTENT_SUFFIXES[3:6]
SUBJECT_GRADE_SUFFIXES = CONTENT_SUFFIXES[6:12]
MATH_GRADE_SUFFIXES = CONTENT_SUFFIXES[6:9]
ENGLISH_GRADE_SUFFIXES = CONTENT_SUFFIXES[9:12]
GRADE_SUBJECT_SUFFIXES = {
    GRADE_SUFFIXES[0]: [CONTENT_SUFFIXES[6], CONTENT_SUFFIXES[9]],
    GRADE_SUFFIXES[1]: [CONTENT_SUFFIXES[7], CONTENT_SUFFIXES[10]],
    GRADE_SUFFIXES[2]: [CONTENT_SUFFIXES[8], CONTENT_SUFFIXES[11]],
}
SUBJECT_TO_GRADE_SUFFIXES = {
    MATH_SUFFIX: MATH_GRADE_SUFFIXES,
    ENGLISH_SUFFIX: ENGLISH_GRADE_SUFFIXES,
}
VOID_TAGS = {"area", "base", "br", "col", "embed", "hr", "img", "input", "link", "meta", "param", "source", "track", "wbr"}
STYLE_CSS = """
:root{color-scheme:light;--bg:#f5f7fb;--surface:#fff;--ink:#182230;--muted:#596579;--line:#dce4ef;--accent:#2364c9;--soft:#eef5ff}
*{box-sizing:border-box}
html,body{overflow-x:hidden}
body{margin:0;font-family:'Noto Sans KR',Arial,sans-serif;line-height:1.75;background:var(--bg);color:var(--ink);word-break:keep-all}
a{color:inherit}
img{display:block;max-width:100%;height:auto;margin-left:auto;margin-right:auto}
.container{width:min(1320px,calc(100% - 2rem));margin:0 auto}
.site-header{border-bottom:1px solid var(--line);background:rgba(255,255,255,.94)}
.nav-wrap,.footer-wrap{display:flex;justify-content:space-between;align-items:center;min-height:64px;gap:1rem}
.brand{font-weight:800;color:var(--ink);text-decoration:none;font-size:1.05rem}
.site-nav{display:flex;gap:.45rem;flex-wrap:wrap}
.site-nav a{padding:.45rem .65rem;border-radius:6px;color:#3d4b63;text-decoration:none;font-size:.94rem}
.site-nav a:hover{background:var(--soft);color:var(--accent)}
.page-layout{padding:2.25rem 0 3rem}
.page-card{width:100%;max-width:1180px;margin:0 auto;background:var(--surface);border:1px solid var(--line);border-radius:8px;padding:clamp(1.35rem,3vw,2.35rem);box-shadow:0 12px 30px rgba(24,34,48,.06)}
.breadcrumb{display:flex;flex-wrap:wrap;gap:.35rem;margin:0 0 1.2rem;color:#748093;font-size:.82rem}
.breadcrumb span{display:flex;align-items:center;gap:.35rem}
.breadcrumb span+span:before{content:'>';color:#a8b2c1}
.breadcrumb a{text-decoration:none}
.eyebrow{margin:0 0 .45rem;font-size:.78rem;font-weight:800;color:var(--accent);text-transform:uppercase;letter-spacing:0}
h1{margin:.2rem 0 .85rem;font-size:clamp(1.75rem,5vw,2.65rem);line-height:1.22;letter-spacing:0}
.fixed-images{display:grid;gap:24px;margin:24px auto 32px}
.page-fixed-image{display:block;width:100%;max-width:860px;height:auto;margin:24px auto 32px;border-radius:18px}
.lead{margin:0 0 1.7rem;padding:1.85rem 2rem;border-left:4px solid var(--accent);border-radius:6px;background:var(--soft);color:#41506a;font-size:1.125rem;line-height:1.9}
.prose{max-width:860px;margin:0 auto;color:#28364a}
.prose h2,.prose h3{line-height:1.35;letter-spacing:0;color:#172033}
.prose h2{margin:2.1rem 0 .8rem;padding-top:.3rem;font-size:clamp(1.3rem,3.5vw,1.65rem);border-top:1px solid var(--line)}
.prose h3{margin:1.5rem 0 .55rem;font-size:1.16rem}
.prose p{margin:0 0 1.15rem;color:#48566d;font-size:1.125rem;line-height:1.9}
.prose a{color:var(--accent);text-underline-offset:3px;text-decoration-thickness:1px}
.prose a:hover{color:#184a99;background:var(--soft)}
.prose blockquote{margin:1.4rem 0;padding:1rem 1.1rem;border-left:4px solid var(--accent);border-radius:8px;background:#f8fbff;color:#34445d}
.prose blockquote p:last-child{margin-bottom:0}
.prose ul{display:grid;gap:.55rem;margin:1rem 0 1.35rem;padding:0;list-style:none}
.prose li{position:relative;padding:.72rem .85rem .72rem 2.25rem;border:1px solid var(--line);border-radius:8px;background:#fbfdff;color:#40506a}
.prose li:before{content:'✓';position:absolute;left:.85rem;top:.72rem;color:var(--accent);font-weight:800}
.prose table{width:100%;margin:1.4rem 0;border-collapse:separate;border-spacing:0;border:1px solid var(--line);border-radius:8px;overflow:hidden;background:#fff;font-size:.95rem}
.prose th,.prose td{padding:.75rem .85rem;border-bottom:1px solid var(--line);vertical-align:top;text-align:left}
.prose th{background:#f3f7fd;color:#26374f;font-weight:800}
.prose tr:last-child th,.prose tr:last-child td{border-bottom:0}
.prose code{padding:.14rem .32rem;border:1px solid #d5dfec;border-radius:5px;background:#f5f7fb;color:#22324a;font-family:Consolas,'Courier New',monospace;font-size:.92em}
.prose pre{margin:1.3rem 0;padding:1rem;border-radius:8px;background:#172033;color:#eef4ff;overflow:auto;line-height:1.65}
.prose pre code{padding:0;border:0;background:transparent;color:inherit}
.faq,.prose details{margin:1rem 0;border:1px solid var(--line);border-radius:8px;background:#fff}
.faq-question,.prose summary{padding:.9rem 1rem;font-weight:800;color:#24344d;cursor:pointer}
.faq-answer,.prose details>*:not(summary){padding:0 1rem 1rem;color:#4a5870}
.button,.btn,.prose .button,.prose .btn{display:inline-flex;align-items:center;justify-content:center;min-height:42px;margin:.25rem .25rem .25rem 0;padding:.65rem .95rem;border:1px solid #a9c7f4;border-radius:8px;background:var(--accent);color:#fff;text-decoration:none;font-weight:800}
.button:hover,.btn:hover,.prose .button:hover,.prose .btn:hover{background:#184a99;color:#fff}
.related-links{max-width:860px;margin:2rem auto 0;padding-top:1.2rem;border-top:1px solid var(--line)}
.related-links h2{margin:0 0 .85rem;font-size:1.05rem;line-height:1.35}
.related-section{margin:1rem 0 0}
.related-section h3{margin:0 0 .55rem;font-size:.92rem;line-height:1.35;color:#596579;font-weight:800}
.related-section+.related-section{margin-top:1.2rem}
.related-links ul{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:.65rem;margin:0;padding:0;list-style:none}
.related-links a{display:block;min-height:44px;padding:.72rem .82rem;border:1px solid var(--line);border-radius:8px;background:#fff;color:#26374f;text-decoration:none;font-weight:700;font-size:.94rem;line-height:1.35}
.related-links a:hover{border-color:#a9c7f4;background:var(--soft);color:var(--accent)}
.home-page .page-layout{padding:2.6rem 0 3.4rem}
.home-page .page-card{max-width:1180px;padding:clamp(1.4rem,3vw,2.6rem)}
.home-hero{padding:0 0 1.35rem}
.home-hero-media{margin:0 0 1.4rem}
.home-hero-copy{min-width:0}
.home-hero h2{margin:.25rem 0 1rem;font-size:1.42rem;line-height:1.5;color:#34445d;font-weight:800}
.home-hero .lead{margin-bottom:0}
.home-hero-image{display:block;width:100%;max-width:1320px;height:auto;margin:0 auto;border-radius:20px;object-fit:cover;box-shadow:0 18px 42px rgba(24,34,48,.14)}
.home-section{margin:1.35rem 0 0;padding:1.2rem;border:1px solid var(--line);border-radius:8px;background:#fbfdff}
.home-section h2{margin:0 0 .45rem;font-size:1.24rem;line-height:1.35}
.home-section p{margin:.35rem 0 1rem;color:#48566d;font-size:1rem;line-height:1.8}
.home-section-intro{max-width:760px}
.home-link-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:.75rem;margin:0;padding:0;list-style:none}
.home-link-grid a{display:block;min-height:50px;padding:.9rem .95rem;border:1px solid var(--line);border-radius:8px;background:#fff;color:#26374f;text-decoration:none;font-weight:800;font-size:.96rem;line-height:1.35;box-shadow:0 4px 12px rgba(24,34,48,.04)}
.home-link-grid a:hover{border-color:#a9c7f4;background:var(--soft);color:var(--accent);transform:translateY(-1px)}
.home-region-wrap{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:1rem}
.home-region-group{margin:0;padding:1rem;border:1px solid var(--line);border-radius:8px;background:#fff}
.home-region-group h3{margin:0 0 .7rem;font-size:1rem;line-height:1.35;color:#26374f}
.home-region-group .home-link-grid{grid-template-columns:1fr;gap:.55rem}
.home-copy{background:#fff}
.home-copy p{max-width:820px;margin:.8rem 0 0;color:#42516a}
.site-footer{border-top:1px solid var(--line);background:#fff;color:#66758c}
@media(max-width:800px){.container{width:100%;max-width:none;padding-left:16px;padding-right:16px}.nav-wrap{align-items:flex-start;flex-direction:column;padding:.75rem 16px}.page-layout{width:100%;padding:1rem 0 2rem}.page-card{width:100%;max-width:none;padding:20px 16px;border-radius:8px}.fixed-images{gap:18px;margin:18px auto 24px}.page-fixed-image{width:100%;max-width:100%;border-radius:14px;margin:0 auto}.related-links ul,.home-link-grid,.home-region-wrap{grid-template-columns:1fr}.home-hero-image{width:100%;border-radius:12px}.prose{max-width:none}.prose table{display:block;overflow-x:auto;white-space:nowrap}.button,.btn,.prose .button,.prose .btn{width:100%}}
@media(max-width:640px){.container{width:100%;max-width:none;padding-left:14px;padding-right:14px}.page-layout{width:100%;padding:.85rem 0 1.7rem}.page-card{width:100%;max-width:none;padding:20px 16px;border-radius:8px}.nav-wrap{gap:.6rem;min-height:56px}.site-nav a{padding:.38rem .52rem;font-size:.9rem}.home-hero{padding:0 0 .9rem}.home-hero-media{margin:0 0 1.1rem}.home-hero-image{width:100%;max-height:250px;object-fit:cover;border-radius:14px}.home-hero h2{font-size:1.34rem;line-height:1.35}.home-section{margin:1rem 0 0;padding:1rem}.home-section h2,.prose h2{font-size:1.75rem;line-height:1.3}.home-region-group h3,.prose h3{font-size:1.3rem;line-height:1.32}.home-section p,.home-copy p,.prose p,.prose li,.related-links a{font-size:1.0625rem;line-height:1.85}.lead{max-width:100%;padding:22px;border-left-width:4px;font-size:1.0625rem;line-height:1.85}.home-link-grid{grid-template-columns:repeat(2,minmax(0,1fr));gap:10px}.home-region-group .home-link-grid{grid-template-columns:1fr}.home-link-grid a{min-height:44px;padding:.68rem .72rem;font-size:.95rem;line-height:1.35}.fixed-images{gap:16px;margin:20px 0}.page-fixed-image{display:block;width:100%;max-width:100%;height:auto;margin:20px auto;border-radius:14px}.related-links{max-width:none;margin-top:1.5rem}.prose ul{gap:.45rem}.prose li{padding:.65rem .75rem .65rem 2rem}}
@media(max-width:420px){.container{padding-left:12px;padding-right:12px}.home-link-grid{grid-template-columns:1fr}h1{font-size:2.375rem;line-height:1.25}.home-section h2,.prose h2{font-size:1.68rem}}
""".strip()


@dataclass
class RegionRow:
    city: str
    district: str
    dong: str


class TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        self.parts.append(data)


class TagChecker(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.stack: list[str] = []
        self.errors: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag not in VOID_TAGS:
            self.stack.append(tag)

    def handle_endtag(self, tag: str) -> None:
        if tag in VOID_TAGS:
            return
        if not self.stack:
            self.errors.append(f"unexpected closing tag: {tag}")
            return
        if self.stack[-1] == tag:
            self.stack.pop()
            return
        self.errors.append(f"mismatched closing tag: {tag}")

    def close(self) -> None:
        super().close()
        if self.stack:
            self.errors.append(f"unclosed tags: {', '.join(self.stack)}")


def compact_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def text_from_html(value: Any) -> str:
    parser = TextExtractor()
    parser.feed(str(value or ""))
    return compact_text(" ".join(parser.parts))


def html_errors(value: Any) -> list[str]:
    html = str(value or "").strip()
    if not html:
        return ["empty_body"]
    checker = TagChecker()
    checker.feed(html)
    checker.close()
    if not re.search(r"<h[1-4]\b[^>]*>.*?</h[1-4]>", html, flags=re.I | re.S):
        checker.errors.append("missing_heading")
    if not re.search(r"<p\b[^>]*>.*?</p>", html, flags=re.I | re.S):
        checker.errors.append("missing_paragraph")
    return checker.errors


def normalize_body(value: Any) -> str:
    return compact_text(str(value or ""))


def slugify_region(value: str) -> str:
    slug = re.sub(r"\s+", "-", value.strip().lower())
    slug = re.sub(r"[^\w가-힣-]+", "", slug)
    return slug.strip("-") or "region"


def keyword_slug(keyword: str) -> str:
    return keyword.strip()


def keyword_title(keyword: str) -> str:
    area = keyword.strip()[:-2] if keyword.strip().endswith("과외") else keyword.strip()
    grade = next((token for token in ("초등", "중등", "고등") if token in area), "")
    if grade:
        area = area.replace(grade, "")
    subject = next((token for token in ("수학", "영어") if token in area), "")
    if subject:
        area = area.replace(subject, "")
    pieces = [area, "과외", "학생", "맞춤수업"] if not grade and not subject else [area, grade, subject, "맞춤수업"]
    return f"{' '.join(piece for piece in pieces if piece)} | {SITE_NAME}"


def keyword_meta(keyword: str) -> dict[str, str]:
    area = keyword.strip()[:-2] if keyword.strip().endswith("과외") else keyword.strip()
    grade = next((token for token in ("초등", "중등", "고등") if token in area), "학생")
    subject = next((token for token in ("수학", "영어") if token in area), "전 과목")
    for token in ("초등", "중등", "고등", "수학", "영어"):
        area = area.replace(token, "")
    return {"area": area or keyword, "grade": grade, "subject": subject}


def generated_body(keyword: str) -> str:
    meta = keyword_meta(keyword)
    area, grade, subject = meta["area"], meta["grade"], meta["subject"]
    return "\n".join(
        [
            f"<p>{escape(keyword)}는 {escape(area)} 생활권의 학교 진도와 학생의 현재 이해도를 함께 보며 학습 방향을 정리해야 하는 주제입니다. 같은 지역과 학년이라도 개념을 놓친 지점, 문제 풀이 속도, 복습 습관이 다르기 때문에 한 가지 방식으로만 판단하기 어렵습니다.</p>",
            f"<h2>{escape(area)} 학습 점검</h2>",
            f"<p>{escape(subject)} 학습에서는 시험 범위, 교과서 진도, 오답이 반복되는 단원, 평소 공부 시간을 함께 확인해야 합니다. 특히 {escape(grade)} 단계에서는 문제를 많이 푸는 것보다 왜 틀렸는지 설명하고 다시 적용하는 과정이 중요합니다.</p>",
            f"<h2>{escape(keyword)} 수업 흐름</h2>",
            "<p>수업은 설명을 듣는 시간, 직접 풀어보는 시간, 풀이 과정을 말로 정리하는 시간을 나누어 운영하는 편이 안정적입니다. 이렇게 하면 학생이 알고 있다고 느낀 부분과 실제로 사용할 수 있는 부분의 차이를 확인할 수 있습니다.</p>",
            f"<p>{escape(area)}에서 {escape(keyword)}를 준비할 때는 한 주 동안 실제로 공부할 수 있는 시간을 먼저 적고, 예습과 복습을 나누어 넣는 것이 좋습니다. 학교 수업을 들은 날에는 핵심 개념을 짧게 정리하고, 다음 날에는 같은 내용을 문제로 확인하며, 주말에는 틀린 유형을 다시 묶어 보는 흐름이 도움이 됩니다.</p>",
            f"<p>마지막으로 {escape(grade)} 학생에게 맞는 {escape(subject)} 학습은 한 번에 큰 변화를 기대하기보다 작은 확인을 꾸준히 반복하는 방향이 효과적입니다. 수업 후에는 그날 배운 내용을 요약하고 다음 학습 전에 같은 유형을 다시 풀어 보며 이해가 유지되는지 점검해야 합니다.</p>",
        ]
    )


def stable_pick(items: list[Any] | tuple[Any, ...], key: str, salt: str) -> Any:
    digest = hashlib.sha256(f"{key}:{salt}".encode("utf-8")).hexdigest()
    return items[int(digest[:8], 16) % len(items)]


def is_elementary_english_keyword(keyword: str) -> bool:
    return keyword.strip().endswith("초등영어과외")


def elementary_english_body(keyword: str) -> str:
    area = keyword.strip()[: -len("초등영어과외")]
    profile = stable_pick(
        [
            {
                "angle": "학생 관찰형",
                "start": "아이의 영어 실력은 단어장 점수만으로 설명되기 어렵습니다.",
                "focus": "수업 시간에 들은 표현을 따라 말하는지, 짧은 문장을 읽을 때 어디에서 멈추는지, 모르는 단어를 만났을 때 바로 포기하지 않는지를 함께 봐야 합니다.",
                "routine": "짧은 낭독, 그림이나 상황을 떠올리는 질문, 오늘 배운 표현을 한 문장으로 다시 말하는 과정을 연결하면 부담이 줄어듭니다.",
            },
            {
                "angle": "학부모 고민형",
                "start": "초등 영어는 시작 시기보다 아이가 영어를 대하는 감정과 습관을 살피는 일이 먼저입니다.",
                "focus": "숙제는 해 오지만 문장을 소리 내어 읽기 어려워하거나, 단어는 외웠는데 뜻을 문장 안에서 연결하지 못하는 경우가 많습니다.",
                "routine": "가정에서는 긴 학습량보다 정해진 시간에 읽고, 표시하고, 짧게 말해보는 흐름을 반복하는 편이 안정적입니다.",
            },
            {
                "angle": "학교생활 적응형",
                "start": "학교 영어 수업에 자연스럽게 참여하려면 교과서 표현을 따라가는 힘과 발표 부담을 낮추는 연습이 필요합니다.",
                "focus": "알파벳과 파닉스가 익숙해 보여도 듣기, 읽기, 쓰기 활동이 함께 나오면 속도가 느려지는 학생이 있습니다.",
                "routine": "수업 전에는 핵심 단어를 가볍게 확인하고, 수업 후에는 교과서 문장을 자기 말로 바꾸어 보는 방식이 도움이 됩니다.",
            },
            {
                "angle": "독서 낭독형",
                "start": "초등 영어에서 낭독은 단순히 크게 읽는 활동이 아니라 문장 구조와 의미를 동시에 확인하는 과정입니다.",
                "focus": "소리와 철자의 연결, 문장의 끊어 읽기, 자주 나오는 표현을 알아보는 경험이 쌓이면 리딩 자신감이 조금씩 생깁니다.",
                "routine": "짧은 지문을 여러 번 읽되 매번 같은 방식으로 반복하지 않고, 그림 설명과 질문 답하기를 섞어야 지루함이 줄어듭니다.",
            },
            {
                "angle": "숙제 루틴형",
                "start": "영어 숙제가 밀리는 아이는 실력이 부족해서라기보다 과제를 시작하는 순서가 정리되지 않은 경우도 많습니다.",
                "focus": "단어 쓰기, 문장 읽기, 듣기 과제, 워크북 풀이가 뒤섞이면 무엇부터 해야 하는지 몰라 시간이 길어집니다.",
                "routine": "먼저 읽고, 모르는 단어를 표시하고, 짧은 문제를 풀고, 마지막에 한 번 더 소리 내어 읽는 순서를 고정하면 부담이 줄어듭니다.",
            },
            {
                "angle": "파닉스 점검형",
                "start": "파닉스는 한 번 끝내는 단원이 아니라 새 단어를 만날 때 계속 확인해야 하는 기초입니다.",
                "focus": "비슷한 소리를 구분하지 못하거나 철자와 발음을 따로 외우는 학생은 쉬운 문장에서도 읽기 속도가 흔들릴 수 있습니다.",
                "routine": "소리 규칙을 짧게 확인한 뒤 실제 단어와 문장 속에서 다시 적용해보면 암기식 학습보다 오래 남습니다.",
            },
            {
                "angle": "단어 누적형",
                "start": "초등 영어 단어는 많이 외우는 것보다 자주 쓰는 표현과 함께 누적되는지가 중요합니다.",
                "focus": "뜻만 외운 단어는 독해나 말하기에서 바로 떠오르지 않기 때문에 그림, 상황, 짧은 문장 안에서 반복해야 합니다.",
                "routine": "하루 학습량을 작게 나누고, 전날 단어를 새 문장에 다시 넣어보는 방식으로 누적하면 잊어버리는 속도를 늦출 수 있습니다.",
            },
            {
                "angle": "문장 읽기형",
                "start": "문장 읽기가 어려운 학생은 단어를 몰라서만이 아니라 어디에서 끊고 어떻게 의미를 묶어야 하는지 모를 때가 많습니다.",
                "focus": "짧은 주어와 동사를 찾고, 반복되는 표현을 표시하고, 문장의 장면을 떠올리는 연습이 함께 필요합니다.",
                "routine": "처음부터 긴 지문을 요구하기보다 두세 문장을 정확히 읽고 자기 말로 설명하는 시간을 만드는 편이 좋습니다.",
            },
            {
                "angle": "자신감 회복형",
                "start": "영어를 싫어한다고 말하는 아이도 작은 성공 경험이 생기면 태도가 달라질 수 있습니다.",
                "focus": "틀린 답을 바로 지적하기보다 어떤 소리와 단어를 이미 알고 있는지 확인하고, 맞힌 부분을 기준으로 다음 활동을 이어가야 합니다.",
                "routine": "짧은 문장을 읽어낸 경험, 스스로 단어를 고른 경험, 질문에 한 문장으로 답한 경험을 기록하면 자신감 회복에 도움이 됩니다.",
            },
            {
                "angle": "생활 리듬형",
                "start": "초등 영어는 생활 리듬 안에 들어와야 오래 이어집니다.",
                "focus": "학원이나 학교 숙제만 몰아서 처리하면 듣기와 읽기 감각이 끊기기 쉬워 매일 짧게 접하는 구조가 필요합니다.",
                "routine": "등교 전 단어 확인, 하교 후 낭독, 주말 복습처럼 시간대를 나누면 아이가 영어를 특별한 부담으로만 느끼지 않습니다.",
            },
        ],
        keyword,
        "profile",
    )
    concern = stable_pick(
        [
            "파닉스가 끝났다고 생각했지만 낯선 단어를 만나면 읽기가 느려지는지",
            "단어는 외우지만 문장 안에서 뜻을 자연스럽게 연결하지 못하는지",
            "교과서 활동은 따라가지만 발표나 말하기 차례가 오면 긴장하는지",
            "숙제를 오래 붙잡고 있어도 실제로 기억하는 표현이 많지 않은지",
            "읽기는 가능한데 내용을 자기 말로 설명하는 데 어려움이 있는지",
            "영어를 틀릴까 봐 피하려는 태도가 먼저 나오는지",
            "듣기에서는 알아듣지만 쓰기나 문장 만들기에서 막히는지",
            "학습량은 적지 않은데 복습 간격이 길어 금방 잊어버리는지",
        ],
        keyword,
        "concern",
    )
    check_items = [
        stable_pick(
            [
                "소리와 철자를 연결해 새 단어를 읽어보는 시간",
                "교과서 표현을 짧은 대화로 바꾸어 말하는 시간",
                "그날 배운 단어를 그림이나 상황과 함께 떠올리는 시간",
            ],
            keyword,
            "check1",
        ),
        stable_pick(
            [
                "낭독할 때 멈추는 지점을 표시하고 다시 읽는 과정",
                "숙제 전후에 모르는 단어를 스스로 구분하는 과정",
                "짧은 문장을 읽고 뜻을 한국어로만 옮기지 않는 과정",
            ],
            keyword,
            "check2",
        ),
        stable_pick(
            [
                "틀린 문제를 바로 지우기보다 왜 헷갈렸는지 말해보는 습관",
                "전날 표현을 새 문장에 넣어보며 기억을 확인하는 습관",
                "읽기, 듣기, 쓰기를 한 번에 몰지 않고 짧게 나누는 습관",
            ],
            keyword,
            "check3",
        ),
    ]
    order = stable_pick(
        [
            ("관찰", "점검", "루틴"),
            ("학교", "습관", "읽기"),
            ("고민", "반응", "다음 단계"),
            ("리듬", "노출", "문장"),
            ("기초", "성공 경험", "복습"),
        ],
        keyword,
        "order",
    )
    return "\n".join(
        [
            f"<p>{escape(keyword)}를 살펴볼 때는 아이가 실제로 영어를 만나는 장면을 구체적으로 보는 것이 좋습니다. {escape(area)} 생활권에서는 학교 수업, 숙제 시간, 이동 동선이 모두 다르기 때문에 같은 초등영어라도 필요한 점검 순서가 달라질 수 있습니다. {profile['start']} {profile['focus']}</p>",
            f"<h2>{escape(area)} 초등영어 학습 점검</h2>",
            f"<p>먼저 확인할 부분은 {concern}입니다. 이 지점이 정리되지 않으면 단어를 많이 외워도 문장 읽기와 듣기 활동에서 다시 막힐 수 있습니다. {escape(area)}초등영어과외를 비교할 때도 교재 이름보다 아이가 어느 활동에서 멈추는지, 어떤 방식으로 설명하면 다시 시도하는지를 관찰하는 편이 현실적입니다.</p>",
            f"<h2>{order[0]}에서 시작하는 수업 흐름</h2>",
            f"<p>{profile['routine']} 초등 단계에서는 한 번에 긴 지문을 끝내는 것보다 짧은 성공 경험을 자주 만드는 편이 좋습니다. 예를 들어 단어를 확인한 뒤 바로 문제를 푸는 방식만 반복하기보다, 읽고 말하고 표시하고 다시 읽는 순서를 아이의 속도에 맞게 조절해야 합니다.</p>",
            "<ul>"
            + "".join(f"<li>{escape(item)}</li>" for item in check_items)
            + "</ul>",
            f"<h2>{order[1]}과 {order[2]}을 연결하는 방법</h2>",
            f"<p>영어가 낯선 아이에게는 매일의 양보다 반복 방식이 더 중요합니다. {escape(area)}에서 학습 시간을 정할 때도 평일에는 짧은 낭독과 단어 확인을 중심으로 두고, 주말에는 지난 표현을 문장 안에서 다시 사용해보는 식으로 나누면 부담이 덜합니다. 이렇게 하면 숙제, 리딩, 듣기 활동이 따로 흩어지지 않고 하나의 루틴으로 이어집니다.</p>",
            f"<p>마지막으로 초등영어는 결과를 서두르기보다 아이가 영어를 피하지 않게 만드는 과정이 필요합니다. {escape(keyword)}를 준비한다면 현재 점수만 보지 말고 파닉스, 어휘, 문장 읽기, 낭독 자신감, 숙제 습관을 차례로 살펴보는 것이 좋습니다. 작은 확인과 복습이 쌓이면 학교 영어 수업에 참여하는 태도도 조금씩 안정될 수 있습니다.</p>",
        ]
    )


def ensure_minimum_body(keyword: str, body: str) -> str:
    if len(text_from_html(body)) >= 800:
        return body
    meta = keyword_meta(keyword)
    area, grade, subject = meta["area"], meta["grade"], meta["subject"]
    extra = (
        f"<p>{escape(keyword)} 페이지에서는 지역명만 바꾸어 정보를 나열하기보다 "
        f"{escape(area)} 생활권에서 실제로 확인해야 할 학습 흐름을 차분히 살펴보는 것이 중요합니다. "
        f"{escape(grade)} 단계의 {escape(subject)} 학습은 현재 이해도, 학교 진도, 복습 간격, "
        "문제 풀이 습관이 함께 맞물릴 때 안정됩니다. 따라서 처음에는 부족한 단원을 크게 판단하기보다 "
        "최근에 틀린 문제, 설명을 듣고 다시 풀 수 있는 문제, 혼자서는 막히는 문제를 나누어 보는 과정이 필요합니다. "
        "이렇게 정리하면 수업 방향이 막연해지지 않고, 다음 학습에서 무엇을 먼저 확인해야 하는지도 분명해집니다.</p>"
    )
    return f"{body}\n{extra}"


STYLE_ANGLES = [
    "학생 관찰형", "부모 고민형", "학교 적응형", "시험 준비형", "생활 습관형",
    "독서형", "오답 분석형", "루틴형", "시간관리형", "심리형",
    "성장 사례형", "비교형", "체크포인트형", "계절형", "방학형",
    "학년 전환형", "내신 중심형", "수행평가형", "진로 연결형", "학습 환경형",
    "복습 설계형", "기초 점검형", "집중력 관리형", "목표 설정형",
]

TITLE_PATTERNS = [
    "{keyword} 학습 정보", "{keyword} 공부 가이드", "{keyword} 학습 환경",
    "{keyword} 교육 정보", "{keyword} 내신 준비", "{keyword} 공부 방향",
    "{keyword} 학습 전략", "{keyword} 학습 리포트", "{keyword} 교육 가이드",
    "{keyword} 과목별 준비", "{keyword} 학년별 안내", "{keyword} 공부 습관",
    "{keyword} 학교생활 학습", "{keyword} 시험 대비 흐름", "{keyword} 기초 점검",
    "{keyword} 학습 계획", "{keyword} 지역 학습 안내", "{keyword} 단계별 정리",
]

DESCRIPTION_PATTERNS = [
    "{area} 생활권에서 {grade} {subject} 학습을 살펴볼 때 필요한 학교 진도, 복습 흐름, 공부 습관을 정리했습니다. 과목별 준비 방향과 지역 학습 환경을 함께 확인해 보세요.",
    "{keyword}를 찾는 학생과 학부모를 위해 {area}의 통학 리듬, 시험 준비, 과목별 기초 점검 포인트를 차분히 정리했습니다.",
    "{area}에서 {subject} 학습 방향을 잡을 때 확인하면 좋은 {grade} 단계의 공부 습관, 오답 관리, 학교생활 흐름을 안내합니다.",
    "{keyword} 페이지에서는 지역 생활권과 학년 특성을 함께 보며 내신 준비, 기초 보완, 복습 루틴을 어떻게 세울지 살펴봅니다.",
    "{area} 학생에게 필요한 {grade} {subject} 학습 관점을 학교 진도, 과제 습관, 시험 대비 흐름으로 나누어 정리했습니다.",
    "{keyword}와 관련해 통학 시간, 공부 공간, 학습 태도, 과목별 보완 방향을 함께 확인할 수 있도록 구성했습니다.",
    "{area}의 학습 환경을 기준으로 {grade} 학생이 {subject}를 준비할 때 놓치기 쉬운 부분과 점검 순서를 정리했습니다.",
    "{keyword}를 준비하기 전 확인하면 좋은 현재 이해도, 오답 습관, 학교 시험 흐름, 생활 리듬을 정보형으로 안내합니다.",
    "{area} 지역의 학교생활과 공부 시간을 함께 고려해 {grade} {subject} 학습 계획을 세우는 데 필요한 내용을 담았습니다.",
    "{keyword}를 알아보는 과정에서 필요한 과목별 기초, 학년별 변화, 내신 대비 흐름을 지역 학습 환경과 함께 정리했습니다.",
    "{area}에서 공부 방향을 고를 때 살펴볼 만한 {subject} 학습 포인트와 {grade} 학생의 복습 습관을 중심으로 안내합니다.",
    "{keyword} 페이지는 학생의 현재 수준, 학교 진도, 생활권의 학습 리듬을 함께 보며 준비 방향을 찾을 수 있게 구성했습니다.",
    "{area} 생활권의 공부 환경을 바탕으로 {grade} {subject} 학습에서 중요한 개념 점검, 과제 관리, 시험 대비 흐름을 정리했습니다.",
    "{keyword}와 관련한 학습 환경, 과목별 준비 방식, 학교생활 속 공부 습관을 한곳에서 차분히 살펴볼 수 있습니다.",
    "{area} 학생이 {subject} 공부를 이어갈 때 필요한 {grade} 단계의 목표 설정, 복습 방식, 실전 점검 포인트를 안내합니다.",
]

INTRO_PATTERNS = [
    "{keyword}는 {area} 생활권의 {grade} {subject} 학습을 차분히 살펴보는 안내입니다.",
    "{area}에서 {grade} {subject} 공부를 준비할 때는 학교 진도와 생활 리듬을 함께 봐야 합니다.",
    "{keyword}를 알아보는 과정에서는 점수보다 먼저 학생이 어디에서 멈추는지 확인하는 일이 중요합니다.",
    "{area} 학생의 {subject} 학습은 통학 시간, 과제량, 복습 습관에 따라 다른 흐름을 보입니다.",
    "{grade} 단계의 {subject} 공부는 현재 이해도와 시험 준비 방식이 함께 맞물릴 때 안정됩니다.",
    "{keyword}는 단순한 과목 선택보다 학생의 하루 공부 흐름을 점검하는 데서 출발합니다.",
    "{area} 생활권에서는 학교마다 과제와 평가 방식이 달라 학습 계획도 세밀하게 볼 필요가 있습니다.",
    "{subject} 성취가 흔들릴 때는 문제 수보다 풀이 과정과 복습 간격을 먼저 살펴야 합니다.",
    "{keyword}를 찾는 가정이라면 학생의 수업 반응과 혼자 공부하는 시간을 함께 확인해 볼 만합니다.",
    "{area}에서 공부 방향을 잡을 때는 가까운 생활권의 학습 분위기와 학교 일정을 같이 고려해야 합니다.",
    "{grade} 학생에게 필요한 {subject} 학습은 한 번의 설명보다 반복 가능한 습관에서 힘을 얻습니다.",
    "{keyword}는 지역 학습 환경과 학생 개인의 속도를 함께 보며 준비 방향을 잡는 주제입니다.",
    "{area}의 학교생활 안에서 {subject} 공부가 어떻게 이어지는지 살피면 필요한 보완점이 더 분명해집니다.",
    "{grade} 과정은 과목별 난도가 바뀌는 시기라 현재 단원 이해도와 생활 습관을 함께 점검해야 합니다.",
    "{keyword}를 검토할 때는 학생이 설명을 들은 뒤 스스로 다시 해낼 수 있는지를 봐야 합니다.",
    "{area}에서 학습 계획을 세울 때는 공부 공간, 이동 시간, 시험 일정이 모두 영향을 줍니다.",
    "{subject} 학습은 기초 확인, 적용 연습, 오답 정리가 균형을 이룰 때 오래 유지됩니다.",
    "{keyword}의 핵심은 많은 내용을 빠르게 넣는 것이 아니라 필요한 순서를 찾는 데 있습니다.",
    "{grade} 학생은 학교 수업을 따라가는 힘과 집에서 정리하는 힘이 함께 필요합니다.",
    "{area} 생활권의 학습 환경을 보면 학생에게 맞는 공부 시간과 복습 방식을 더 현실적으로 잡을 수 있습니다.",
    "{keyword}에서는 학생의 현재 상태를 과목, 학년, 생활 리듬으로 나누어 보는 관점이 필요합니다.",
    "{subject} 공부가 막힐 때는 어려운 단원을 늘리기보다 작은 실수의 원인을 찾는 편이 좋습니다.",
    "{area}의 학교와 가정 사이에서 공부 시간이 어떻게 쓰이는지 보면 보완 방향이 달라집니다.",
    "{grade} 과정의 학습은 시험 직전만이 아니라 평소 정리 습관에서 차이가 생깁니다.",
    "{keyword}를 준비한다면 학생이 이해한 내용과 실제로 써먹는 내용의 차이를 살펴야 합니다.",
    "{area} 학생에게 맞는 학습 방향은 학교 일정과 개인 속도를 함께 고려할 때 더 선명해집니다.",
    "{subject_topic} 개념, 표현, 문제 풀이가 따로 움직이지 않도록 연결하는 과정이 중요합니다.",
    "{keyword}는 지역 생활권의 분위기와 학생의 학습 습관을 함께 살펴볼 때 의미가 커집니다.",
    "{grade} 학습은 작은 성공 경험을 쌓으며 다음 단원으로 넘어가는 흐름이 필요합니다.",
    "{area}에서 {subject}를 준비하는 학생이라면 공부량보다 공부 순서를 먼저 점검해 볼 필요가 있습니다.",
    "{keyword}는 학교생활, 과목 특성, 가정 학습 루틴을 함께 정리하는 데 도움이 됩니다.",
    "{grade} {subject} 공부는 당장의 문제 풀이와 장기적인 습관 사이의 균형을 잡아야 합니다.",
]

H2_TOPICS = [
    ("학습 환경", "{area} 학습 환경 살펴보기"),
    ("현재 수준", "현재 이해도와 출발점"),
    ("학교 흐름", "학교 진도와 평가 흐름"),
    ("공부 습관", "공부 습관을 잡는 방식"),
    ("오답 관리", "오답을 다루는 기준"),
    ("시간 계획", "통학과 공부 시간 조율"),
    ("과목 전략", "{subject} 학습 전략"),
    ("학년 변화", "{grade} 단계에서 달라지는 점"),
    ("생활 리듬", "생활 리듬과 복습 루틴"),
    ("점검 기준", "가정에서 확인할 체크포인트"),
    ("시험 준비", "시험 전후 준비 흐름"),
    ("성장 기록", "작은 변화 기록하기"),
]

FAQ_POOL = [
    "{area}에서 {keyword}를 알아볼 때 먼저 볼 부분은 무엇인가요?",
    "{grade} 학생은 {subject} 공부를 어느 순서로 시작하면 좋을까요?",
    "학교 진도와 개인 복습은 어떻게 맞추는 것이 좋나요?",
    "오답 정리는 매일 해야 하나요, 주 단위로 해도 괜찮나요?",
    "단원 이해가 약한 학생은 어떤 방식으로 점검하면 좋나요?",
    "시험 직전에는 새 문제와 복습 중 무엇을 우선해야 하나요?",
    "공부 시간이 짧은 학생도 루틴을 만들 수 있나요?",
    "{area} 생활권의 통학 시간이 학습 계획에 영향을 주나요?",
    "학부모가 집에서 확인할 수 있는 신호는 무엇인가요?",
    "기초가 약한 학생은 교재보다 어떤 부분을 먼저 봐야 하나요?",
    "수행평가 준비는 평소 공부와 어떻게 연결하면 좋나요?",
    "{subject} 공부에서 실수가 반복될 때는 어떻게 해야 하나요?",
    "학년이 바뀌기 전 어떤 내용을 점검하면 좋나요?",
    "방학에는 선행과 복습 중 무엇을 더 봐야 하나요?",
    "주말 학습은 평일 복습과 어떻게 다르게 잡으면 좋나요?",
    "학교 시험 범위가 넓을 때 계획은 어떻게 나누면 좋나요?",
    "집중 시간이 짧은 학생에게 필요한 방식은 무엇인가요?",
    "문제를 많이 풀어도 변화가 적을 때 확인할 점은 무엇인가요?",
    "개념 설명을 이해했는데 시험에서 틀리는 이유는 무엇인가요?",
    "가정 학습 시간을 무리 없이 늘리는 방법은 무엇인가요?",
    "학습 기록은 어느 정도까지 남기는 것이 좋나요?",
    "내신 준비와 장기 실력 관리는 어떻게 나누면 좋나요?",
    "학교별 평가 방식 차이를 어떻게 확인하면 좋나요?",
    "학생이 부담을 느낄 때 학습량을 줄여도 괜찮나요?",
    "짧은 복습이 실제로 도움이 되는 경우는 언제인가요?",
    "학습 공간을 바꾸면 집중력에 차이가 생길 수 있나요?",
    "시험 이후 틀린 문제는 언제 다시 보는 것이 좋나요?",
    "진로 목표가 아직 없어도 학습 계획이 필요한가요?",
    "학년 전환기에는 어떤 습관을 먼저 잡아야 하나요?",
    "과목 흥미가 낮은 학생은 어디서부터 시작하면 좋나요?",
    "성적보다 먼저 살펴야 할 학습 태도는 무엇인가요?",
    "공부 계획이 자주 무너질 때 원인은 무엇일까요?",
    "학교 숙제와 별도 복습은 어떻게 구분하면 좋나요?",
    "개념 노트가 모든 학생에게 필요한가요?",
    "틀린 문제를 다시 풀 때 답을 외우는 문제는 어떻게 줄이나요?",
    "평일 학습량과 주말 학습량의 균형은 어떻게 잡나요?",
    "중간고사와 기말고사 사이에는 무엇을 준비하면 좋나요?",
    "학습 속도가 느린 학생에게 필요한 기준은 무엇인가요?",
    "학생이 스스로 설명하는 시간을 왜 넣어야 하나요?",
    "문제집을 바꾸기 전에 확인할 점은 무엇인가요?",
    "학교 수업을 잘 듣는 학생도 별도 정리가 필요한가요?",
    "복습 간격은 어느 정도가 적당한가요?",
    "학습 목표는 크게 잡는 편이 좋나요, 작게 나누는 편이 좋나요?",
    "공부 습관이 약할 때 가장 먼저 바꿀 행동은 무엇인가요?",
    "부모가 매일 확인하면 오히려 부담이 될 수 있나요?",
    "시험 불안이 있는 학생은 어떤 준비가 필요할까요?",
    "풀이 속도가 느릴 때는 시간을 재는 연습부터 해야 하나요?",
    "서술형 답안은 어떻게 연습하면 좋나요?",
    "독해형 문제는 어떤 순서로 접근하면 좋나요?",
    "계산 실수나 표현 실수가 반복될 때 공통 원인이 있나요?",
    "학습 공백이 있는 학생은 전 학년 내용을 얼마나 봐야 하나요?",
    "수업 후 바로 복습하지 못하면 효과가 많이 줄어드나요?",
    "스스로 공부하는 시간을 늘리려면 어떤 장치가 필요할까요?",
    "지역 학습 분위기를 확인할 때 어떤 정보를 보면 좋나요?",
    "도서관이나 스터디 공간 활용은 언제 도움이 되나요?",
    "학교 행사 기간에는 학습 계획을 어떻게 줄이면 좋나요?",
    "방학 직후에는 어떤 방식으로 페이스를 회복하면 좋나요?",
    "새 학기 초반에는 어떤 과목 신호를 먼저 봐야 하나요?",
    "학습 방향이 맞는지 확인하는 가장 현실적인 방법은 무엇인가요?",
    "학생이 질문을 잘 하지 않을 때는 어떻게 접근하면 좋나요?",
    "성장 과정을 기록하면 어떤 점이 달라지나요?",
    "짧은 성공 경험을 만드는 방법은 무엇인가요?",
]

CHECKLIST_POOL = [
    "최근 학교 진도와 학생이 어려워한 단원을 함께 적어본다.",
    "틀린 문제를 계산 실수, 개념 혼동, 시간 부족으로 나누어 본다.",
    "평일 공부 가능 시간과 실제 집중 시간을 따로 확인한다.",
    "통학 후 바로 공부할 수 있는 날과 휴식이 필요한 날을 구분한다.",
    "시험 범위가 나오기 전에도 매주 복습할 단원을 정해 둔다.",
    "설명을 들은 뒤 학생이 자기 말로 다시 말할 수 있는지 본다.",
    "숙제 완료 여부보다 어떤 문제에서 오래 멈췄는지 확인한다.",
    "주말에는 새 단원보다 지난 주 내용을 다시 묶어 본다.",
    "학교 과제와 별도 복습을 한 목록에 섞지 않는다.",
    "학습 공간의 소음, 이동 거리, 시작 시간을 함께 점검한다.",
    "하루 목표는 문제 수보다 완료 가능한 행동으로 적는다.",
    "오답은 바로 지우지 말고 처음 생각한 이유를 남긴다.",
    "단원별로 쉬운 문제와 어려운 문제를 한 번씩 섞어 본다.",
    "수행평가 일정은 시험 계획과 같은 달력에 표시한다.",
    "학생이 부담을 말하는 표현을 기록해 학습량 조절에 반영한다.",
    "방학 계획은 선행, 복습, 생활 리듬 세 영역으로 나눈다.",
    "새 학기 전에는 전 학기에서 반복된 실수부터 확인한다.",
    "짧은 낭독이나 설명 시간을 넣어 이해 여부를 확인한다.",
    "문제집을 바꾸기 전 현재 교재의 오답 유형을 먼저 본다.",
    "학교 시험 후에는 점수보다 틀린 단원 분포를 살핀다.",
    "집중 시간이 짧다면 20분 단위로 학습 행동을 쪼갠다.",
    "도서관이나 공부 공간을 이용할 때 이동 시간을 계획에 넣는다.",
    "부모 확인은 매일 잔소리보다 주간 점검표로 가볍게 둔다.",
    "스스로 질문한 내용을 따로 적어 다음 학습의 출발점으로 삼는다.",
    "개념 노트는 길게 쓰기보다 헷갈린 조건을 짧게 정리한다.",
    "시험 2주 전에는 새 계획보다 기존 복습 간격을 좁힌다.",
    "생활 리듬이 흔들린 주에는 학습량보다 시작 시간을 회복한다.",
    "진로 관심이 있는 학생은 관련 단원과 학습 목표를 연결해 본다.",
    "성취가 작아도 스스로 해낸 부분을 기록한다.",
    "풀이 속도 연습은 정확도가 안정된 뒤에 시간을 재며 진행한다.",
    "서술형은 정답보다 조건을 빠뜨리지 않는지 먼저 본다.",
    "독해형 문제는 핵심 문장 표시와 근거 찾기를 함께 연습한다.",
    "계산 실수는 반복되는 위치와 숫자 처리 습관을 따로 본다.",
    "단어 암기는 뜻 확인 뒤 문장 속 사용으로 이어간다.",
    "학습 공백은 전 범위를 훑기보다 자주 막히는 단원부터 좁힌다.",
    "수업 직후 10분 정리를 습관으로 만들 수 있는지 확인한다.",
    "학생이 선택할 수 있는 작은 목표를 하나 남긴다.",
    "학교 행사나 평가 기간에는 계획을 줄여 지속성을 우선한다.",
    "가정 학습이 길어질 때는 휴식 시간을 먼저 정해 둔다.",
    "학습 변화는 하루보다 2주 단위로 비교한다.",
    "같은 실수가 세 번 반복되면 풀이 순서를 다시 점검한다.",
    "과목별 어려움을 한 문장으로 설명해 보게 한다.",
    "학습량이 늘어도 피로가 누적되지 않는지 살핀다.",
    "수업 자료와 학교 프린트가 따로 놀지 않게 묶어 둔다.",
    "목표는 성적만이 아니라 행동 변화도 함께 적는다.",
]

CONCLUSION_PATTERNS = [
    "{keyword}를 준비할 때는 빠른 결과보다 학생이 다시 해낼 수 있는 흐름을 만드는 일이 중요합니다.",
    "{area} 생활권의 학습 환경을 함께 보면 {grade} {subject} 준비 방향을 더 현실적으로 잡을 수 있습니다.",
    "{subject} 공부는 작은 점검을 꾸준히 이어갈 때 시험 전 부담이 줄어듭니다.",
    "{keyword}는 학생의 현재 위치를 차분히 확인하고 다음 단계를 정하는 과정으로 바라보는 편이 좋습니다.",
    "{grade} 학습은 무리한 양보다 반복 가능한 루틴을 만드는 데서 안정감을 얻습니다.",
    "{area}에서 공부 계획을 세운다면 학교 일정과 가정 학습 시간을 함께 조율해 보는 것이 필요합니다.",
    "{subject}의 변화는 한 번의 설명보다 오답을 다시 다루는 태도에서 시작되는 경우가 많습니다.",
    "{keyword}를 살펴보는 과정이 학생에게 맞는 공부 속도와 기준을 찾는 계기가 되면 좋습니다.",
    "{grade} 단계에서는 작은 성취를 기록하며 다음 단원으로 넘어가는 흐름이 중요합니다.",
    "{area}의 생활 리듬을 고려한 계획은 학습을 오래 이어가는 데 도움이 됩니다.",
    "{subject} 학습은 기초, 적용, 복습이 따로 움직이지 않도록 연결하는 과정이 필요합니다.",
    "{keyword}에서는 학생의 학교생활과 과목별 어려움을 함께 보는 관점이 중요합니다.",
    "{grade} 학생에게 맞는 방향은 현재 부족한 부분을 정확히 좁히는 데서 출발합니다.",
    "{area}에서 학습 환경을 정리하면 공부 시간과 복습 방식을 더 구체적으로 잡을 수 있습니다.",
    "{subject} 준비는 시험 직전보다 평소 정리 습관에서 차이가 만들어집니다.",
    "{keyword}는 학습량을 늘리기 전에 학생에게 맞는 순서를 찾는 데 의미가 있습니다.",
    "{grade} 과정의 공부는 자신감과 실전 감각을 함께 키우는 방향이 좋습니다.",
    "{area} 생활권의 학교 흐름을 살피면 내신 준비와 일상 복습의 균형을 맞추기 쉽습니다.",
    "{subject} 학습은 틀린 문제를 대하는 방식이 달라질 때 지속적인 변화가 생깁니다.",
    "{keyword}를 통해 학생의 학습 습관을 한 번 더 점검하고 필요한 준비를 차분히 이어가길 바랍니다.",
]

REGION_FEATURES = [
    "학교와 집 사이의 이동 시간이 공부 시작 시간에 영향을 주는 생활권",
    "도서관이나 조용한 공부 공간을 활용하기 쉬운 환경",
    "학교별 과제량과 평가 일정이 학습 리듬을 바꾸는 지역",
    "학원가 접근성과 가정 학습 시간이 함께 고려되는 생활권",
    "주말 복습 시간을 확보하기 좋은 동선이 있는 지역",
    "등하교 후 휴식과 공부 시작 시간을 조율해야 하는 환경",
    "시험 기간에 학교 프린트와 교과서 정리가 중요해지는 분위기",
    "초등, 중등, 고등 생활권이 가까워 학년 전환 흐름을 보기 좋은 지역",
    "가정 학습과 외부 공부 공간을 함께 활용할 수 있는 환경",
    "학교 행사와 수행평가 일정이 학습 계획에 영향을 주는 생활권",
]


def stable_sample(items: list[Any] | tuple[Any, ...], key: str, salt: str, count: int) -> list[Any]:
    scored = []
    for index, item in enumerate(items):
        digest = hashlib.sha256(f"{key}:{salt}:{index}".encode("utf-8")).hexdigest()
        scored.append((digest, item))
    return [item for _, item in sorted(scored)[:count]]


def quality_meta(keyword: str) -> dict[str, str]:
    raw = keyword.strip()
    base = raw[:-2] if raw.endswith("과외") else raw
    grade = next((token for token in ("초등", "중등", "고등") if token in base), "")
    subject = next((token for token in ("수학", "영어") if token in base), "")
    area = base
    for token in ("초등", "중등", "고등", "수학", "영어"):
        area = area.replace(token, "")
    subject_value = subject or "전 과목"
    subject_topic = f"{subject_value}{'은' if subject_value in {'수학', '전 과목'} else '는'}"
    return {
        "keyword": raw,
        "area": area or raw,
        "grade": grade or "학생",
        "subject": subject_value,
        "subject_topic": subject_topic,
        "type": f"{grade or '지역'}{subject or '기본'}",
    }


def render_pattern(pattern: str, meta: dict[str, str]) -> str:
    return pattern.format(**meta)


def quality_title(keyword: str) -> str:
    meta = quality_meta(keyword)
    pattern = stable_pick(TITLE_PATTERNS, keyword, "title")
    return f"{render_pattern(pattern, meta)} | {SITE_NAME}"


def quality_description(keyword: str) -> str:
    meta = quality_meta(keyword)
    text = render_pattern(stable_pick(DESCRIPTION_PATTERNS, keyword, "description"), meta)
    if len(text) < 120:
        text = (
            f"{text} {meta['area']}의 학교생활, 통학 리듬, 가정 학습 시간을 함께 고려해 "
            f"{meta['grade']} {meta['subject']} 준비 방향을 더 구체적으로 살펴볼 수 있습니다."
        )
    if len(text) > 180:
        text = text[:177].rstrip() + "..."
    return text


def quality_body(keyword: str) -> str:
    meta = quality_meta(keyword)
    area = meta["area"]
    subject = meta["subject"]
    grade = meta["grade"]
    angle = stable_pick(STYLE_ANGLES, keyword, "angle")
    feature = stable_pick(REGION_FEATURES, keyword, "region-feature")
    intro = render_pattern(stable_pick(INTRO_PATTERNS, keyword, "intro"), meta)
    h2_items = stable_sample(H2_TOPICS, keyword, "h2", 5)
    checklist = [render_pattern(item, meta) for item in stable_sample(CHECKLIST_POOL, keyword, "checklist", 5)]
    faqs = [render_pattern(item, meta) for item in stable_sample(FAQ_POOL, keyword, "faq", 5)]
    conclusion = render_pattern(stable_pick(CONCLUSION_PATTERNS, keyword, "conclusion"), meta)
    type_note = {
        "초등영어": "소리 내어 읽기, 짧은 문장 이해, 학교 영어 활동 적응을 무리 없이 연결하는 관점이 필요합니다.",
        "중등영어": "어휘, 문법, 독해, 서술형이 함께 움직이므로 시험 범위와 평소 복습을 분리하지 않는 편이 좋습니다.",
        "고등영어": "내신 지문 분석과 모의고사 독해가 동시에 필요해 시간 배분과 근거 찾기 습관이 중요합니다.",
        "초등수학": "계산 정확도, 개념 언어, 문제를 읽는 습관을 함께 다루어야 다음 단원으로 넘어가기 쉽습니다.",
        "중등수학": "개념 이해와 유형 적용 사이의 간격을 줄이고, 서술형 풀이 과정을 정리하는 시간이 필요합니다.",
        "고등수학": "단원별 개념 연결과 시험 시간 안배가 함께 요구되므로 오답 원인을 깊게 보는 방식이 중요합니다.",
    }.get(meta["type"], "과목과 학년이 섞여 있는 지역 페이지에서는 학생의 생활 리듬과 현재 학습 습관을 함께 보는 관점이 필요합니다.")
    paragraphs = [
        f"<p>{escape(keyword)}의 첫 점검은 {escape(area)} 생활권과 {escape(grade)} {escape(subject)} 학습 흐름을 함께 보는 데서 시작합니다. {escape(intro)} {escape(angle)} 관점으로 보면 {escape(area)}의 {escape(feature)}이라는 점을 함께 고려하게 됩니다. 학생마다 이해 속도와 복습 방식이 다르기 때문에 처음부터 많은 내용을 넣기보다 현재 막히는 장면을 작게 나누어 살피는 편이 좋습니다.</p>",
        f"<p>{escape(type_note)} 특히 {escape(grade)} 단계에서는 학교 수업을 따라가는 힘과 집에서 다시 정리하는 힘이 함께 필요합니다. {escape(subject)} 학습이 흔들릴 때는 문제 수, 공부 시간, 교재 이름보다 어떤 상황에서 실수가 반복되는지 먼저 확인해야 합니다.</p>",
    ]
    section_bodies = {
        "학습 환경": f"{area}에서는 통학 후 컨디션, 저녁 공부 시작 시간, 주말 복습 가능 여부가 학습 지속성에 영향을 줍니다. 가까운 도서관이나 조용한 공부 공간을 활용할 수 있다면 짧은 복습 시간을 고정하기가 조금 수월합니다.",
        "현재 수준": f"현재 수준은 점수 하나로만 보기 어렵습니다. 설명을 들으면 이해하지만 혼자 풀 때 막히는지, 쉬운 문제에서 실수가 잦은지, 새로운 유형을 만났을 때 어떤 순서로 접근하는지를 나누어 봐야 합니다.",
        "학교 흐름": f"학교 진도는 교과서 단원, 프린트, 수행평가 일정이 함께 움직입니다. 시험 범위가 나오기 전에도 최근 수업에서 반복된 개념과 과제 유형을 정리해 두면 시험 직전의 부담이 줄어듭니다.",
        "공부 습관": f"공부 습관은 긴 계획표보다 실제로 지킬 수 있는 행동에서 시작됩니다. 하루에 한 단원을 끝내기 어렵다면 짧은 확인, 핵심 문제, 오답 표시처럼 작게 나누어 반복하는 방식이 더 안정적입니다.",
        "오답 관리": f"오답 관리는 틀린 문제를 다시 푸는 일로 끝나지 않습니다. 왜 그렇게 생각했는지, 어느 조건을 놓쳤는지, 같은 실수가 어느 단원에서 반복되는지를 기록해야 다음 학습의 기준이 생깁니다.",
        "시간 계획": f"시간 계획은 공부량을 늘리는 약속이 아니라 시작 시간을 지키는 장치에 가깝습니다. 통학, 식사, 휴식 시간을 감안해 짧은 복습 블록을 먼저 확보하면 평일에도 학습 흐름을 유지하기 쉽습니다.",
        "과목 전략": f"{subject} 학습에서는 기초 확인과 적용 연습의 균형이 중요합니다. 이미 아는 내용을 빠르게 넘기기보다 헷갈리는 조건을 짚고, 비슷한 문제에서 다시 적용해 보는 시간이 필요합니다.",
        "학년 변화": f"{grade} 과정에서는 이전 학년의 습관이 그대로 이어지지 않을 수 있습니다. 평가 방식, 과제량, 수업 속도가 달라지는 지점을 미리 확인하면 학년 전환기의 흔들림을 줄일 수 있습니다.",
        "생활 리듬": f"생활 리듬은 학습 결과에 생각보다 큰 영향을 줍니다. 피곤한 날에는 새 내용을 무리하게 늘리기보다 짧은 복습과 정리 중심으로 조정해야 다음 날 공부 흐름이 끊기지 않습니다.",
        "점검 기준": f"가정에서는 정답 개수보다 학생이 설명할 수 있는 범위를 확인하는 편이 좋습니다. 스스로 말로 정리하지 못하는 부분은 아직 적용이 흔들릴 가능성이 있으므로 다시 다루어야 합니다.",
        "시험 준비": f"시험 준비는 범위가 확정된 뒤 시작하면 늦어질 수 있습니다. 평소에는 단원별 핵심 개념과 자주 틀리는 유형을 모아 두고, 시험 전에는 그 목록을 중심으로 반복하는 방식이 효율적입니다.",
        "성장 기록": f"작은 변화도 기록해 두면 학습 방향을 조정하기 쉽습니다. 풀이 시간이 줄었는지, 설명이 더 정확해졌는지, 같은 실수가 줄었는지를 보면 단순한 점수보다 구체적인 변화가 보입니다.",
    }
    for key, heading in h2_items:
        paragraphs.append(f"<h2>{escape(render_pattern(heading, meta))}</h2>")
        paragraphs.append(f"<p>{escape(section_bodies[key])}</p>")
    paragraphs.append("<h2>체크리스트</h2>")
    paragraphs.append("<ul>" + "".join(f"<li>{escape(item)}</li>" for item in checklist) + "</ul>")
    paragraphs.append("<h2>자주 묻는 질문</h2>")
    for question in faqs:
        answer = f"{area}의 학교생활과 {grade} {subject} 학습 흐름을 함께 보며 학생이 실제로 지킬 수 있는 기준부터 정하는 것이 좋습니다. 한 번에 많은 항목을 바꾸기보다 현재 가장 자주 막히는 장면을 정하고, 그 부분을 짧게 반복해 확인하는 방식이 안정적입니다."
        paragraphs.append(f"<details><summary>{escape(question)}</summary><p>{escape(answer)}</p></details>")
    paragraphs.append(f"<p>{escape(conclusion)} 지역의 공부 환경, 학교 일정, 학생의 속도를 함께 살피면 준비 과정이 더 구체적이고 부담이 적어집니다.</p>")
    return ensure_quality_length(keyword, "\n".join(paragraphs))


def ensure_quality_length(keyword: str, body: str) -> str:
    if len(text_from_html(body)) >= 1200:
        return body
    meta = quality_meta(keyword)
    extra = (
        f"<p>{escape(meta['area'])}에서 학습 계획을 이어갈 때는 주간 단위의 작은 점검이 도움이 됩니다. "
        f"{escape(meta['grade'])} {escape(meta['subject'])} 공부는 하루의 결과만으로 판단하기보다 최근 2주 동안 반복된 실수, "
        "공부를 시작한 시간, 복습을 마친 뒤 스스로 설명할 수 있었던 내용을 함께 보아야 합니다. "
        "이런 기준을 두면 학습량을 무리하게 늘리지 않아도 필요한 보완점을 찾기 쉽습니다.</p>"
    )
    return f"{body}\n{extra}"


def inspect_workbooks() -> tuple[list[RegionRow], list[dict[str, Any]]]:
    print("READ_CHECK")
    print(f"REGION_FILE: {REGION_XLSX} exists={REGION_XLSX.exists()} size={REGION_XLSX.stat().st_size if REGION_XLSX.exists() else None}")
    print(f"CONTENT_FILE: {CONTENT_XLSX} exists={CONTENT_XLSX.exists()} size={CONTENT_XLSX.stat().st_size if CONTENT_XLSX.exists() else None}")
    if not REGION_XLSX.exists() or not CONTENT_XLSX.exists():
        raise SystemExit("STOP: required latest Excel files are missing")

    regions: list[RegionRow] = []
    region_wb = load_workbook(REGION_XLSX, read_only=True, data_only=True)
    print("\nREGION_SHEET_ROWS")
    for ws in region_wb.worksheets:
        rows = 0
        for values in ws.iter_rows(values_only=True):
            vals = [(str(v).strip() if v is not None else "") for v in values[:3]]
            if len(vals) >= 3 and vals[0] and vals[1] and vals[2] and vals[0] != "시도":
                rows += 1
                regions.append(RegionRow(vals[0], vals[1], vals[2]))
        print(f"{ws.title}: {rows}")
    region_wb.close()

    content_rows: list[dict[str, Any]] = []
    content_wb = load_workbook(CONTENT_XLSX, read_only=True, data_only=True)
    print("\nCONTENT_SHEET_ROWS")
    for ws in content_wb.worksheets:
        rows = 0
        for row_number, values in enumerate(ws.iter_rows(values_only=True), start=1):
            keyword = str(values[0]).strip() if values and values[0] is not None else ""
            body = str(values[1]) if len(values) > 1 and values[1] is not None else ""
            if not keyword or keyword == "키워드":
                continue
            rows += 1
            content_rows.append({"sheet": ws.title, "row": row_number, "keyword": keyword, "body": body})
        print(f"{ws.title}: {rows}")
    content_wb.close()

    print("\nTOTALS")
    print(f"region_rows: {len(regions)}")
    print(f"content_rows: {len(content_rows)}")
    if len(content_rows) != EXPECTED_CONTENT_ROWS:
        raise SystemExit(f"STOP: content row count is {len(content_rows)}, expected {EXPECTED_CONTENT_ROWS}. Generation aborted.")
    return regions, content_rows


def build_content_pages(rows: list[dict[str, Any]]):
    duplicate_keywords = {key for key, count in Counter(row["keyword"] for row in rows).items() if count > 1}
    duplicate_bodies = {body for body, count in Counter(normalize_body(row["body"]) for row in rows if normalize_body(row["body"])).items() if count > 1}
    pages = []
    duplicate_pages = []
    repaired_pages = []
    failed_pages = []
    seen = set()
    for row in rows:
        keyword = row["keyword"]
        body = row["body"]
        reasons = []
        text_len = len(text_from_html(body))
        if not normalize_body(body):
            reasons.append("empty_body")
        if text_len < MIN_TEXT_LENGTH:
            reasons.append(f"short_body:{text_len}")
        reasons.extend(html_errors(body))
        if normalize_body(body) in duplicate_bodies:
            reasons.append("duplicate_body")
        if keyword in duplicate_keywords:
            duplicate_pages.append({**row, "slug": keyword_slug(keyword), "reason": "duplicate_keyword"})
        if keyword in seen:
            continue
        seen.add(keyword)
        try:
            final_body = quality_body(keyword)
            if reasons:
                repaired_pages.append({**row, "slug": keyword_slug(keyword), "reasons": sorted(set(reasons))})
            pages.append(
                {
                    "type": "content",
                    "sheet": row["sheet"],
                    "row": row["row"],
                    "slug": keyword_slug(keyword),
                    "keyword": keyword,
                    "h1": keyword,
                    "title": quality_title(keyword),
                    "description": quality_description(keyword),
                    "content": final_body,
                }
            )
        except Exception as exc:
            failed_pages.append({**row, "slug": keyword_slug(keyword), "error": str(exc)})
    return pages, duplicate_pages, repaired_pages, failed_pages


def append_link(links: list[tuple[str, str]], page: dict[str, Any] | None, current_slug: str = "") -> None:
    if not page or page["slug"] == current_slug:
        return
    href = f"/{page['slug'].strip('/')}/"
    if href not in {item[1] for item in links}:
        links.append((page["keyword"], href))


def content_page_for_area(content_by_keyword: dict[str, dict[str, Any]], area: str, suffix: str = BASE_SUFFIX) -> dict[str, Any] | None:
    return content_by_keyword.get(f"{area}{suffix}")


def content_pages_for_areas(
    content_by_keyword: dict[str, dict[str, Any]],
    areas: list[str] | set[str],
    suffix: str = BASE_SUFFIX,
) -> list[dict[str, Any] | None]:
    return [content_page_for_area(content_by_keyword, area, suffix) for area in areas]


def content_pages_for_suffixes(
    content_by_keyword: dict[str, dict[str, Any]],
    area: str,
    suffixes: list[str],
) -> list[dict[str, Any] | None]:
    return [content_page_for_area(content_by_keyword, area, suffix) for suffix in suffixes]


def suffix_priority_for(current_suffix: str = BASE_SUFFIX) -> list[str]:
    if current_suffix in MATH_GRADE_SUFFIXES or current_suffix == MATH_SUFFIX:
        return [BASE_SUFFIX, *GRADE_SUFFIXES, MATH_SUFFIX, *MATH_GRADE_SUFFIXES, ENGLISH_SUFFIX, *ENGLISH_GRADE_SUFFIXES]
    if current_suffix in ENGLISH_GRADE_SUFFIXES or current_suffix == ENGLISH_SUFFIX:
        return [BASE_SUFFIX, *GRADE_SUFFIXES, ENGLISH_SUFFIX, *ENGLISH_GRADE_SUFFIXES, MATH_SUFFIX, *MATH_GRADE_SUFFIXES]
    return CONTENT_SUFFIX_PRIORITY


def add_area_variants(
    links: list[tuple[str, str]],
    content_by_keyword: dict[str, dict[str, Any]],
    area: str,
    current_slug: str = "",
    current_suffix: str = BASE_SUFFIX,
) -> None:
    for suffix in suffix_priority_for(current_suffix):
        append_link(links, content_page_for_area(content_by_keyword, area, suffix), current_slug)
        if len(links) >= RELATED_LIMIT:
            return


def suffix_for_keyword(keyword: str, area: str) -> str:
    suffix = keyword[len(area) :]
    return suffix if suffix in CONTENT_SUFFIXES else BASE_SUFFIX


def page_kind_from_suffix(suffix: str) -> str:
    if suffix in SUBJECT_GRADE_SUFFIXES:
        return "subject_grade"
    if suffix in SUBJECT_SUFFIXES:
        return "subject"
    if suffix in GRADE_SUFFIXES:
        return "grade"
    return "base"


def sibling_dongs(by_district: dict[tuple[str, str], list[str]], city: str, district: str, dong: str = "") -> list[str]:
    return [item for item in by_district[(city, district)] if item != dong]


def flatten_related_sections(sections: list[dict[str, Any]]) -> list[tuple[str, str]]:
    links: list[tuple[str, str]] = []
    for section in sections:
        links.extend(section["links"])
    return links


def classify_content_page(page: dict[str, Any], regions: list[RegionRow]) -> dict[str, str] | None:
    keyword = page["keyword"]
    best: dict[str, str] | None = None
    best_len = -1
    for item in regions:
        candidates = [
            ("city", item.city, {"city": item.city, "district": "", "dong": "", "area": item.city}),
            ("district", item.district, {"city": item.city, "district": item.district, "dong": "", "area": item.district}),
            ("dong", item.dong, {"city": item.city, "district": item.district, "dong": item.dong, "area": item.dong}),
        ]
        for level, area, context in candidates:
            if any(keyword == f"{area}{suffix}" for suffix in CONTENT_SUFFIXES) and len(area) > best_len:
                best = {"level": level, "suffix": suffix_for_keyword(keyword, area), **context}
                best_len = len(area)
    return best


def related_links_for_context(
    context: dict[str, str],
    content_by_keyword: dict[str, dict[str, Any]],
    by_city: dict[str, set[str]],
    by_district: dict[tuple[str, str], set[str]],
    current_slug: str = "",
) -> list[tuple[str, str]]:
    links: list[tuple[str, str]] = []
    level = context["level"]
    city = context["city"]
    district = context["district"]
    dong = context["dong"]
    area = context["area"]
    suffix = context.get("suffix", "과외")

    if level == "city":
        add_area_variants(links, content_by_keyword, city, current_slug, suffix)
        for child_district in sorted(by_city[city]):
            append_link(links, content_page_for_area(content_by_keyword, child_district, suffix), current_slug)
            if len(links) >= RELATED_LIMIT:
                return links
        return links[:RELATED_LIMIT]

    if level == "district":
        append_link(links, content_page_for_area(content_by_keyword, city), current_slug)
        append_link(links, content_page_for_area(content_by_keyword, district), current_slug)
        if len(links) < RELATED_LIMIT:
            for child_dong in sorted(by_district[(city, district)]):
                append_link(links, content_page_for_area(content_by_keyword, child_dong, suffix), current_slug)
                if len(links) >= RELATED_LIMIT:
                    return links
        add_area_variants(links, content_by_keyword, district, current_slug, suffix)
        append_link(links, content_page_for_area(content_by_keyword, city, suffix), current_slug)
        return links[:RELATED_LIMIT]

    if level == "dong":
        append_link(links, content_page_for_area(content_by_keyword, city), current_slug)
        append_link(links, content_page_for_area(content_by_keyword, district), current_slug)
        append_link(links, content_page_for_area(content_by_keyword, dong), current_slug)
        add_area_variants(links, content_by_keyword, dong, current_slug, suffix)
        if len(links) < RELATED_LIMIT:
            for sibling_dong in sorted(by_district[(city, district)]):
                if sibling_dong == dong:
                    continue
                append_link(links, content_page_for_area(content_by_keyword, sibling_dong, suffix), current_slug)
                if len(links) >= RELATED_LIMIT:
                    return links
        append_link(links, content_page_for_area(content_by_keyword, district, suffix), current_slug)
        append_link(links, content_page_for_area(content_by_keyword, city, suffix), current_slug)
        return links[:RELATED_LIMIT]

    return links


def related_sections_for_context(
    context: dict[str, str],
    content_by_keyword: dict[str, dict[str, Any]],
    by_city: dict[str, list[str]],
    by_district: dict[tuple[str, str], list[str]],
    current_slug: str = "",
) -> list[dict[str, Any]]:
    sections: list[dict[str, Any]] = []
    seen_hrefs: set[str] = set()
    total_links = 0

    def add_section(title: str, candidates: list[dict[str, Any] | None]) -> None:
        nonlocal total_links
        if total_links >= RELATED_TOTAL_LIMIT:
            return
        links: list[tuple[str, str]] = []
        for page in candidates:
            if total_links >= RELATED_TOTAL_LIMIT or len(links) >= RELATED_SECTION_LIMIT:
                break
            if not page or page["slug"] == current_slug:
                continue
            href = f"/{page['slug'].strip('/')}/"
            if href in seen_hrefs:
                continue
            seen_hrefs.add(href)
            links.append((page["keyword"], href))
            total_links += 1
        if links:
            sections.append({"title": title, "links": links})

    level = context["level"]
    city = context["city"]
    district = context["district"]
    dong = context["dong"]
    suffix = context.get("suffix", BASE_SUFFIX)
    kind = page_kind_from_suffix(suffix)

    if level == "city":
        if kind == "base":
            add_section("하위 구/지역", content_pages_for_areas(content_by_keyword, by_city[city], BASE_SUFFIX))
            add_section("시 단위 과목", content_pages_for_suffixes(content_by_keyword, city, SUBJECT_SUFFIXES))
            add_section("시 단위 학년", content_pages_for_suffixes(content_by_keyword, city, GRADE_SUFFIXES))
        else:
            add_section("상위 지역", [content_page_for_area(content_by_keyword, city, BASE_SUFFIX)])
            add_section("시 단위 관련", content_pages_for_suffixes(content_by_keyword, city, SUBJECT_SUFFIXES + GRADE_SUFFIXES + SUBJECT_GRADE_SUFFIXES))
            add_section("하위 구/지역", content_pages_for_areas(content_by_keyword, by_city[city], suffix))
        return sections

    if level == "district":
        add_section("상위 지역", [content_page_for_area(content_by_keyword, city, BASE_SUFFIX)])
        if kind == "base":
            add_section("하위 동", content_pages_for_areas(content_by_keyword, by_district[(city, district)], BASE_SUFFIX))
            add_section("구 단위 과목", content_pages_for_suffixes(content_by_keyword, district, SUBJECT_SUFFIXES))
            add_section("구 단위 학년", content_pages_for_suffixes(content_by_keyword, district, GRADE_SUFFIXES))
        else:
            add_section("구 단위 관련", content_pages_for_suffixes(content_by_keyword, district, SUBJECT_SUFFIXES + GRADE_SUFFIXES + SUBJECT_GRADE_SUFFIXES))
            add_section("하위 동", content_pages_for_areas(content_by_keyword, by_district[(city, district)], suffix))
            add_section("시 단위 관련", [content_page_for_area(content_by_keyword, city, suffix)])
        return sections

    if level == "dong":
        parent_candidates = [
            content_page_for_area(content_by_keyword, city, BASE_SUFFIX),
            content_page_for_area(content_by_keyword, district, BASE_SUFFIX),
        ]
        if kind != "base":
            parent_candidates.append(content_page_for_area(content_by_keyword, dong, BASE_SUFFIX))
        add_section("상위 지역", parent_candidates)

        if kind == "base":
            add_section("같은 지역 과목", content_pages_for_suffixes(content_by_keyword, dong, SUBJECT_SUFFIXES))
            add_section("같은 지역 학년", content_pages_for_suffixes(content_by_keyword, dong, GRADE_SUFFIXES))
            add_section("같은 지역 세부 과목", content_pages_for_suffixes(content_by_keyword, dong, SUBJECT_GRADE_SUFFIXES))
            add_section("주변 지역", content_pages_for_areas(content_by_keyword, sibling_dongs(by_district, city, district, dong), BASE_SUFFIX))
        elif kind == "subject":
            other_subjects = [item for item in SUBJECT_SUFFIXES if item != suffix]
            subject_label = "수학" if suffix == MATH_SUFFIX else "영어"
            add_section("상위 지역 같은 과목", [
                content_page_for_area(content_by_keyword, district, suffix),
                content_page_for_area(content_by_keyword, city, suffix),
            ])
            add_section("같은 지역 다른 과목", content_pages_for_suffixes(content_by_keyword, dong, other_subjects))
            add_section("같은 지역 학년", content_pages_for_suffixes(content_by_keyword, dong, GRADE_SUFFIXES))
            add_section(f"같은 지역 {subject_label} 세부", content_pages_for_suffixes(content_by_keyword, dong, SUBJECT_TO_GRADE_SUFFIXES.get(suffix, [])))
            add_section("주변 지역", content_pages_for_areas(content_by_keyword, sibling_dongs(by_district, city, district, dong), suffix))
        elif kind == "grade":
            add_section("상위 지역 같은 학년", [
                content_page_for_area(content_by_keyword, district, suffix),
                content_page_for_area(content_by_keyword, city, suffix),
            ])
            add_section("같은 지역 과목", content_pages_for_suffixes(content_by_keyword, dong, SUBJECT_SUFFIXES))
            add_section("같은 지역 세부 과목", content_pages_for_suffixes(content_by_keyword, dong, GRADE_SUBJECT_SUFFIXES.get(suffix, [])))
            add_section("주변 지역", content_pages_for_areas(content_by_keyword, sibling_dongs(by_district, city, district, dong), suffix))
        else:
            grade_suffix = next((item for item in GRADE_SUFFIXES if suffix.startswith(item.replace(BASE_SUFFIX, ""))), "")
            subject_suffix = MATH_SUFFIX if suffix in MATH_GRADE_SUFFIXES else ENGLISH_SUFFIX
            same_grade_detail = [item for item in SUBJECT_GRADE_SUFFIXES if item != suffix and item in GRADE_SUBJECT_SUFFIXES.get(grade_suffix, [])]
            add_section("상위 지역 같은 세부 과목", [
                content_page_for_area(content_by_keyword, district, suffix),
                content_page_for_area(content_by_keyword, city, suffix),
            ])
            add_section("같은 지역 관련", [
                content_page_for_area(content_by_keyword, dong, subject_suffix),
                content_page_for_area(content_by_keyword, dong, grade_suffix),
            ])
            add_section("같은 지역 세부 과목", content_pages_for_suffixes(content_by_keyword, dong, same_grade_detail))
            add_section("주변 지역", content_pages_for_areas(content_by_keyword, sibling_dongs(by_district, city, district, dong), suffix))
        return sections

    return sections


def assign_content_related_links(content_pages: list[dict[str, Any]], regions: list[RegionRow]) -> None:
    by_city = defaultdict(list)
    by_district = defaultdict(list)
    for item in regions:
        if item.district not in by_city[item.city]:
            by_city[item.city].append(item.district)
        if item.dong not in by_district[(item.city, item.district)]:
            by_district[(item.city, item.district)].append(item.dong)

    content_by_keyword = {page["keyword"]: page for page in content_pages}
    for page in content_pages:
        context = classify_content_page(page, regions)
        sections = (
            related_sections_for_context(context, content_by_keyword, by_city, by_district, page["slug"])
            if context
            else []
        )
        page["related_sections"] = sections
        page["related_links"] = flatten_related_sections(sections)


def region_page(slug: str, h1: str, title: str, links: list[tuple[str, str]]):
    link_html = "".join(f'<li><a href="{escape(href)}">{escape(label)}</a></li>' for label, href in links[:RELATED_LIMIT]) or "<li>연결할 세부 페이지가 아직 없습니다.</li>"
    return {"type": "region", "slug": slug, "keyword": h1, "h1": h1, "title": f"{title} | {SITE_NAME}", "description": f"{h1} 기준 지역과 학습 페이지입니다.", "content": f"<p>{escape(h1)} 기준의 지역 구조와 관련 학습 페이지를 확인할 수 있습니다.</p><ul>{link_html}</ul>"}


def build_region_pages(regions: list[RegionRow], content_pages: list[dict[str, Any]]):
    return []


def body_for_output(content: str) -> str:
    content = re.sub(r"<h1\b([^>]*)>", r"<h2\1>", content, flags=re.I)
    return re.sub(r"</h1>", "</h2>", content, flags=re.I)


def top_nav(regions):
    return "".join(f'<a href="/{keyword_slug(f"{city}{BASE_SUFFIX}")}/">{escape(city)}</a>' for city in sorted({r.city for r in regions}))


def absolute_url(path: str = "") -> str:
    clean_path = path.strip("/")
    return f"{BASE_URL}/{clean_path}/" if clean_path else f"{BASE_URL}/"


def fixed_image_name_for_slug(slug: str) -> str:
    digest = hashlib.sha256(slug.strip("/").encode("utf-8")).hexdigest()
    index = int(digest, 16) % len(FIXED_IMAGE_NAMES)
    return FIXED_IMAGE_NAMES[index]


def fixed_image_url(name: str) -> str:
    return f"{FIXED_IMAGE_URL_BASE}/{name}"


def fixed_image_src(name: str) -> str:
    return f"{FIXED_IMAGE_SRC_BASE}/{name}"


def fixed_images_html() -> str:
    images = "".join(
        f'<img class="page-fixed-image" src="{fixed_image_src(name)}" alt="에듀가이드 학습 정보 이미지" width="{FIXED_IMAGE_WIDTH}" height="{FIXED_IMAGE_HEIGHT}" loading="lazy" decoding="async">'
        for name in FIXED_IMAGE_NAMES
    )
    return f'<div class="fixed-images">{images}</div>'


def breadcrumb_items(slug: str, title: str) -> list[dict[str, Any]]:
    items = [{"@type": "ListItem", "position": 1, "name": SITE_NAME, "item": absolute_url()}]
    parts = [part for part in slug.strip("/").split("/") if part]
    for index, part in enumerate(parts, start=2):
        path = "/".join(parts[: index - 1])
        name = title if index == len(parts) + 1 else part
        items.append({"@type": "ListItem", "position": index, "name": name, "item": absolute_url(path)})
    return items


def breadcrumb_html(page: dict[str, Any]) -> str:
    items = breadcrumb_items(page["slug"], page["h1"])
    return "".join(
        f'<span><a href="{escape(item["item"])}">{escape(str(item["name"]))}</a></span>'
        for item in items
    )


def structured_data(page: dict[str, Any], canonical: str) -> str:
    web_page = {
        "@context": "https://schema.org",
        "@type": "WebPage",
        "name": page["title"],
        "url": canonical,
        "description": page["description"],
        "isPartOf": {"@type": "WebSite", "name": SITE_NAME, "url": absolute_url()},
    }
    if page.get("image"):
        web_page["image"] = page["image"]
    graph = [
        {
            "@context": "https://schema.org",
            "@type": "WebSite",
            "name": SITE_NAME,
            "url": absolute_url(),
        },
        web_page,
        {
            "@context": "https://schema.org",
            "@type": "BreadcrumbList",
            "itemListElement": breadcrumb_items(page["slug"], page["h1"]),
        },
    ]
    return "\n".join(
        f'<script type="application/ld+json">{json.dumps(item, ensure_ascii=False)}</script>'
        for item in graph
    )


def render_related_sections(page: dict[str, Any]) -> str:
    sections = page.get("related_sections") or []
    if sections:
        rendered_sections = []
        for section in sections:
            links = "".join(
                f'<li><a href="{escape(href)}">{escape(label)}</a></li>'
                for label, href in section["links"][:RELATED_SECTION_LIMIT]
            )
            if links:
                rendered_sections.append(
                    f'<section class="related-section"><h3>{escape(section["title"])}</h3><ul>{links}</ul></section>'
                )
        return "".join(rendered_sections)

    related_links = page.get("related_links", [])[:RELATED_LIMIT]
    links = "".join(f'<li><a href="{escape(href)}">{escape(label)}</a></li>' for label, href in related_links)
    return f"<ul>{links}</ul>" if links else ""


def render_page(page, all_pages, regions):
    nav = top_nav(regions)
    canonical = absolute_url(page["slug"])
    content = body_for_output(page["content"])
    related = render_related_sections(page)
    selected_image = fixed_image_url(fixed_image_name_for_slug(page["slug"]))
    page = {**page, "image": selected_image}
    crumbs = breadcrumb_html(page)
    json_ld = structured_data(page, canonical)
    fixed_images = fixed_images_html()
    return f'''<!DOCTYPE html><html lang="ko"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><meta name="description" content="{escape(page['description'])}"><title>{escape(page['title'])}</title><link rel="canonical" href="{canonical}"><meta property="og:type" content="website"><meta property="og:site_name" content="{SITE_NAME}"><meta property="og:title" content="{escape(page['title'])}"><meta property="og:description" content="{escape(page['description'])}"><meta property="og:url" content="{canonical}"><meta property="og:image" content="{selected_image}"><meta name="twitter:card" content="summary_large_image"><meta name="twitter:image" content="{selected_image}">{json_ld}<link rel="stylesheet" href="/static/css/style.css"></head><body class="detail-page"><header class="site-header"><div class="container nav-wrap"><a class="brand" href="/">{SITE_NAME}</a><nav class="site-nav">{nav}</nav></div></header><main class="container page-layout"><article class="page-card"><nav class="breadcrumb" aria-label="breadcrumb">{crumbs}</nav><p class="eyebrow">{escape(page['type'])}</p><h1>{escape(page['h1'])}</h1>{fixed_images}<p class="lead">{escape(page['description'])}</p><div class="prose">{content}</div><nav class="related-links"><h2>관련 페이지</h2>{related}</nav></article></main><footer class="site-footer"><div class="container footer-wrap"><p>{SITE_NAME}</p></div></footer></body></html>'''


def render_home(content_pages, region_pages, regions):
    nav = top_nav(regions)
    content_by_keyword = {page["keyword"]: page for page in content_pages}
    cities: list[str] = []
    districts_by_city: dict[str, list[str]] = defaultdict(list)
    dongs_by_district: dict[tuple[str, str], list[str]] = defaultdict(list)
    for item in regions:
        if item.city not in cities:
            cities.append(item.city)
        if item.district not in districts_by_city[item.city]:
            districts_by_city[item.city].append(item.district)
        if item.dong not in dongs_by_district[(item.city, item.district)]:
            dongs_by_district[(item.city, item.district)].append(item.dong)

    def page_for(area: str, suffix: str = BASE_SUFFIX) -> dict[str, Any] | None:
        return content_by_keyword.get(f"{area}{suffix}")

    def page_for_keyword(keyword: str) -> dict[str, Any] | None:
        return content_by_keyword.get(keyword)

    def link_list(pages: list[dict[str, Any] | None], limit: int = 24) -> str:
        links: list[tuple[str, str]] = []
        seen: set[str] = set()
        for page in pages:
            if not page:
                continue
            href = f"/{page['slug'].strip('/')}/"
            if href in seen:
                continue
            seen.add(href)
            links.append((page["keyword"], href))
            if len(links) >= limit:
                break
        return "".join(f'<li><a href="{escape(href)}">{escape(label)}</a></li>' for label, href in links)

    preferred_city_names = ["광주", "울산", "전주", "인천", "대전", "대구"]
    city_pages = [page_for(city) for city in preferred_city_names] + [page_for(city) for city in cities]
    district_pages = [page_for(district) for city in cities for district in districts_by_city[city]]
    major_region_links = link_list(city_pages + district_pages, 8)
    subject_links = link_list(
        [page_for_keyword(keyword) for keyword in ["수학과외", "영어과외"]]
        + [page_for(city, suffix) for city in preferred_city_names + cities for suffix in SUBJECT_SUFFIXES],
        6,
    )
    grade_links = link_list(
        [page_for_keyword(keyword) for keyword in ["초등과외", "중등과외", "고등과외"]]
        + [page_for(city, suffix) for city in preferred_city_names + cities for suffix in GRADE_SUFFIXES],
        6,
    )
    detail_groups = []
    for city in cities[:3]:
        group_pages: list[dict[str, Any] | None] = [page_for(city)]
        for district in districts_by_city[city][:3]:
            group_pages.append(page_for(district))
        links = link_list(group_pages, 4)
        if links:
            detail_groups.append(f'<section class="home-region-group"><h3>{escape(city)} 지역별 탐색</h3><ul class="home-link-grid">{links}</ul></section>')
    detail_html = f'<div class="home-region-wrap">{"".join(detail_groups)}</div>'
    guide_links = link_list(
        [page_for_keyword(keyword) for keyword in [
            "수완동과외",
            "수완동수학과외",
            "수완동고등과외",
            "수완동고등수학과외",
            "광산구과외",
            "광산구수학과외",
            "광주과외",
            "광주수학과외",
            "울산과외",
            "울산수학과외",
            "전주과외",
            "전주영어과외",
        ]]
        + city_pages
        + district_pages,
        12,
    )
    title = f"에듀가이드 지역별 과외 학습 허브 | {SITE_NAME}"
    meta_description = "에듀가이드는 광주, 울산, 전주 지역의 과외 정보를 시·구·동, 과목, 학년 기준으로 정리한 교육 정보 허브입니다. 지역별 학습 환경, 수학·영어 과목 준비, 초등·중등·고등 학년별 공부 흐름을 한눈에 살펴보고 필요한 지역 안내 페이지로 쉽고 빠르게 이동할 수 있습니다."
    description = "에듀가이드는 지역별 과외 정보를 한곳에서 살펴볼 수 있도록 정리한 교육 정보 허브입니다. 광주, 울산, 전주처럼 지역 범위가 다른 페이지를 시·구·동 단위로 나누고, 수학·영어 과목과 초등·중등·고등 학년 흐름을 함께 연결해 필요한 정보를 단계적으로 찾을 수 있게 구성했습니다. 처음 방문한 사용자는 주요 지역에서 출발해 세부 지역과 과목 페이지로 이동하며 학습 환경, 준비 방식, 공부 습관 관련 내용을 비교해 볼 수 있습니다. 각 페이지는 지역명과 학습 주제를 함께 다루기 때문에 내신 준비, 과목별 기초 점검, 학년 전환 시기처럼 상황에 맞는 정보를 차분히 탐색하는 데 도움이 됩니다."
    canonical = absolute_url()
    home_page = {"slug": "", "h1": SITE_NAME, "title": title, "description": meta_description}
    json_ld = structured_data(home_page, canonical)
    return f'''<!DOCTYPE html><html lang="ko"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><meta name="description" content="{escape(meta_description)}"><title>{escape(title)}</title><link rel="canonical" href="{canonical}"><meta property="og:type" content="website"><meta property="og:site_name" content="{SITE_NAME}"><meta property="og:title" content="{escape(title)}"><meta property="og:description" content="{escape(meta_description)}"><meta property="og:url" content="{canonical}">{json_ld}<link rel="stylesheet" href="static/css/style.css"></head><body class="home-page"><header class="site-header"><div class="container nav-wrap"><a class="brand" href="/">{SITE_NAME}</a><nav class="site-nav">{nav}</nav></div></header><main class="container page-layout"><section class="page-card"><section class="home-hero"><div class="home-hero-media"><img class="home-hero-image" src="images/hero/{HERO_IMAGE_NAME}" alt="에듀가이드 지역별 학습 정보" loading="eager" decoding="async" fetchpriority="high" width="{HERO_IMAGE_WIDTH}" height="{HERO_IMAGE_HEIGHT}"></div><div class="home-hero-copy"><p class="eyebrow">지역별 과외 학습 허브</p><h1>에듀가이드</h1><h2>지역별 과외 정보를 한눈에 살펴보는 교육 정보 허브</h2><p class="lead">{escape(description)}</p></div></section><section class="home-section"><h2>주요 지역 바로가기</h2><p class="home-section-intro">먼저 큰 지역 단위에서 출발하면 하위 구와 동 페이지로 자연스럽게 이동할 수 있습니다.</p><ul class="home-link-grid">{major_region_links}</ul></section><section class="home-section"><h2>과목별 바로가기</h2><p class="home-section-intro">수학과 영어처럼 자주 찾는 과목 페이지를 중심으로 지역별 학습 정보를 확인할 수 있습니다.</p><ul class="home-link-grid">{subject_links}</ul></section><section class="home-section"><h2>학년별 바로가기</h2><p class="home-section-intro">초등, 중등, 고등 단계별로 필요한 학습 관점이 다르므로 학년 페이지에서 흐름을 나누어 볼 수 있습니다.</p><ul class="home-link-grid">{grade_links}</ul></section><section class="home-section"><h2>지역별 세부 탐색</h2><p class="home-section-intro">시 단위 페이지와 주요 하위 지역을 묶어 전체 구조를 빠르게 살펴볼 수 있도록 정리했습니다.</p>{detail_html}</section><section class="home-section home-copy"><h2>학습 안내</h2><p>지역별 학습 환경은 학교 밀집도, 통학 동선, 생활권에 따라 다르게 나타납니다. 같은 시 안에서도 구나 동에 따라 자주 찾는 과목과 학년 흐름이 달라질 수 있으므로, 큰 지역 페이지에서 출발해 세부 지역 페이지로 이동하며 정보를 좁혀 보는 방식이 유용합니다.</p><p>초등 학습은 기본 개념과 학습 습관을 만드는 데 초점이 있고, 중등 학습은 과목별 단원 이해와 내신 대비 흐름이 중요합니다. 고등 학습은 학교 시험, 모의고사, 진로 선택과 연결되기 때문에 학년 페이지를 함께 살펴보면 준비 방향을 더 분명하게 잡을 수 있습니다.</p><p>수학은 개념 이해, 문제 풀이 과정, 오답 정리가 이어질 때 학습 흐름이 안정됩니다. 영어는 어휘, 문법, 독해, 듣기처럼 영역이 나뉘기 때문에 현재 부족한 부분을 확인하고 학년별 난이도에 맞게 준비하는 것이 좋습니다.</p><p>에듀가이드의 지역 페이지는 시·구·동, 과목, 학년 조합으로 연결되어 있습니다. 먼저 거주지나 학교 생활권에 가까운 지역을 고른 뒤, 수학·영어 또는 초등·중등·고등 페이지로 이동하면 내신 준비와 공부 습관에 필요한 정보를 순서대로 확인할 수 있습니다.</p></section><section class="home-section"><h2>주요 페이지 바로가기</h2><p class="home-section-intro">자주 확인할 만한 대표 페이지를 적게 모아 메인에서 바로 이동할 수 있게 했습니다.</p><ul class="home-link-grid">{guide_links}</ul></section></section></main><footer class="site-footer"><div class="container footer-wrap"><p>{SITE_NAME}</p></div></footer></body></html>'''


def render_404(regions):
    nav = top_nav(regions)
    region_links = "".join(
        f'<li><a href="/{keyword_slug(f"{city}{BASE_SUFFIX}")}/">{escape(city)}과외</a></li>'
        for city in sorted({item.city for item in regions})[:6]
    )
    return f'''<!DOCTYPE html><html lang="ko"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><meta name="robots" content="noindex,follow"><title>페이지를 찾을 수 없습니다 | {SITE_NAME}</title><link rel="canonical" href="{BASE_URL}/404.html"><link rel="stylesheet" href="static/css/style.css"></head><body class="home-page"><header class="site-header"><div class="container nav-wrap"><a class="brand" href="/">{SITE_NAME}</a><nav class="site-nav">{nav}</nav></div></header><main class="container page-layout"><section class="page-card"><p class="eyebrow">404</p><h1>페이지를 찾을 수 없습니다</h1><p class="lead">요청하신 페이지가 이동되었거나 존재하지 않습니다. 에듀가이드 홈 또는 주요 지역 페이지에서 다시 필요한 학습 정보를 찾아보세요.</p><section class="home-section"><h2>주요 지역 링크</h2><ul class="home-link-grid">{region_links}<li><a href="/">홈으로 이동</a></li></ul></section></section></main><footer class="site-footer"><div class="container footer-wrap"><p>{SITE_NAME}</p></div></footer></body></html>'''


def write_file(path: Path, content: str):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def write_json(path: Path, payload: Any):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def ensure_site_images() -> None:
    for name in FIXED_IMAGE_NAMES:
        image_path = FIXED_IMAGE_DIR / name
        if not image_path.exists():
            raise SystemExit(f"STOP: missing image: {image_path}")
    hero_image_path = HERO_IMAGE_DIR / HERO_IMAGE_NAME
    if not hero_image_path.exists():
        raise SystemExit(f"STOP: missing hero image: {hero_image_path}")


def clean():
    for path in (OUTPUT_DIR, DATA_DIR):
        if path.exists():
            shutil.rmtree(path)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    STATIC_DIR.mkdir(parents=True, exist_ok=True)
    write_file(
        STATIC_DIR / "css" / "style.css",
        STYLE_CSS,
    )
    ensure_site_images()


def main():
    regions, content_rows = inspect_workbooks()
    clean()
    content_pages, duplicate_pages, repaired_pages, failed_pages = build_content_pages(content_rows)
    assign_content_related_links(content_pages, regions)
    region_pages = build_region_pages(regions, content_pages)
    all_pages = content_pages + region_pages
    shutil.copytree(STATIC_DIR, OUTPUT_DIR / "static", dirs_exist_ok=True)
    shutil.copytree(FIXED_IMAGE_DIR, OUTPUT_DIR / "images" / "thumbs" / "fixed", dirs_exist_ok=True)
    shutil.copytree(HERO_IMAGE_DIR, OUTPUT_DIR / "images" / "hero", dirs_exist_ok=True)
    write_file(OUTPUT_DIR / "index.html", render_home(content_pages, region_pages, regions))
    write_file(OUTPUT_DIR / "404.html", render_404(regions))
    for page in all_pages:
        write_file(OUTPUT_DIR / page["slug"] / "index.html", render_page(page, all_pages, regions))
    sitemap_entries = [absolute_url()] + [absolute_url(p["slug"]) for p in all_pages]
    write_file(OUTPUT_DIR / "sitemap.xml", "<?xml version=\"1.0\" encoding=\"UTF-8\"?>\n<urlset xmlns=\"http://www.sitemaps.org/schemas/sitemap/0.9\">\n" + "\n".join(f"  <url><loc>{escape(url)}</loc></url>" for url in sitemap_entries) + "\n</urlset>\n")
    write_file(OUTPUT_DIR / "robots.txt", f"User-agent: *\nAllow: /\nSitemap: {SITE_URL}/sitemap.xml\n")
    payload = {"source_files": {"regions": str(REGION_XLSX), "contents": str(CONTENT_XLSX)}, "counts": {"region_rows": len(regions), "content_rows": len(content_rows), "content_pages": len(content_pages), "region_pages": len(region_pages)}, "pages": all_pages}
    write_json(DATA_DIR / "pages_generated.json", payload)
    write_json(DATA_DIR / "pages.json", payload)
    write_json(DATA_DIR / "duplicate_pages.json", duplicate_pages)
    write_json(DATA_DIR / "failed_pages.json", failed_pages)
    write_json(DATA_DIR / "repaired_pages.json", repaired_pages)
    print("\nGENERATED")
    print(f"content_pages: {len(content_pages)}")
    print(f"region_pages: {len(region_pages)}")
    print(f"repaired: {len(repaired_pages)}")
    print(f"duplicates: {len(duplicate_pages)}")
    print(f"failed: {len(failed_pages)}")


if __name__ == "__main__":
    main()
